from openpyxl import load_workbook
import pandas as pd
import os

wb = load_workbook(filename="SP500_ESG.xlsx") # loading the excel

list_req = ['AAPL', 'MSFT', 'AMZN', 'JNJ', 'WMT', 'JPM', 'UNH', 'NVDA', 'BAC', 'VZ',\
                'CMCSA', 'ADBE', 'NFLX', 'KO', 'NKE', 'MRK', 'PFE', 'INTC',\
                'ORCL', 'ABT', 'CSCO', 'TMO', 'QCOM', 'TMUS'] #list of Ticks



def retrieve_score(wb,tick,increase,write):
    
    """
    This function accesses the Excel file containing the Bloomberg Scores and retrieve them
    then put them in a txt file.
    ==============================================================================================================

    Input:
    ------

        wb: loaded workbook using the openxyl librairy
        tick: ticker of the company
        increase: Boolean: if True: print the change in score accross years
                           if False: print the original score
    
    Output:
    -------

        None
    """
    sheetn = wb.sheetnames[1:] # get all the excel sheets
    sheet_names = ['Environmental','Social','Governance','ESG_combined','ESG_controvesies','ESG_score'] # their names
    
    if write:
        list_year= ["09","10","11","12","13","14",\
                    "15","16","17","18","19","20",\
                    "21"]

    else:
        list_year = [tick+"09",tick+"10",tick+"11",tick+"12",tick+"13",tick+"14",\
                    tick+"15",tick+"16",tick+"17",tick+"18",tick+"19",tick+"20",\
                    tick+"21"] # indeces for the dataframe
    
    sr = wb[sheetn[1]]

    # first find the row of the Ticker
    
    row = None
    wb[sheetn[1]]
    for i in range(1,sr.max_row+1):
        k = 'A' + str(i)
        if sr[k].value == tick:
            row = i
        else:
            continue
    
    assert row != None, "Ticker Not Found"

    ############## Since you guys never answered i made it work for both incrase or just plain scores

    # checking if option selected: for each:
    #
    #            - Check if the approriate folder is present if not create it
    #            - define the cell's column to loop through
    #            - define the possible error list that could be
    #            - create a file name to append at the end

    if increase==True:
        if write==True:
            if  not os.path.isdir(os.getcwd().replace("\\","/") + "/Bloomberg_Score_Change"):
                os.makedirs(os.getcwd().replace("\\","/") + "/Bloomberg_Score_Change")
            filename = os.getcwd().replace("\\","/") + '/Bloomberg_Score_Change'+'/Bsi_'+tick+'.txt'

        list_cell = ['AD'+ str(row),'AE'+ str(row),'AF'+ str(row),'AG'+ str(row),'AH'+ str(row),'AI'+ str(row),\
                    'AJ'+ str(row),'AK'+ str(row),'AL'+ str(row),'AM'+ str(row),'AN'+ str(row),'AO'+ str(row),\
                    'AP'+ str(row)]
        er = ["#DIV/0!","#VALUE!"]
        

    else:
        if write==True:

            if  not os.path.isdir(os.getcwd().replace("\\","/") + "/Bloomberg_Score"):
                os.makedirs(os.getcwd().replace("\\","/") + "/Bloomberg_Score")
            filename =os.getcwd().replace("\\","/") + '/Bloomberg_Score'+'/Bs_'+tick+'.txt'

        list_cell = ['F'+ str(row),'G'+ str(row),'H'+ str(row),'I'+ str(row),'J'+ str(row),'K'+ str(row),\
                    'L'+ str(row),'M'+ str(row),'N'+ str(row),'O'+ str(row),'P'+ str(row),'Q'+ str(row),\
                    'R'+ str(row)]
        er = []
        
    
    
    assert len(list_cell) == len(list_year), "Missing Year or Cells"

    score = [] #initializing the score

    for sheet in sheetn: # looping through all the sheets
        
        sr = wb[sheet] # assigining the searech engine
        
        topic_score = [] # creating the topic score list

        for cell in list_cell: # for every cell in the list cells 

            if sr[cell].value not in er: # if its not in the error list
                
                topic_score.append(sr[cell].value) # append it

            else: # otherwise
                topic_score.append(0) # append 0 (only the case for the change in scores)

        score.append(topic_score) # appending to score list
        del topic_score # deleting to conserve memory
    
    
    data = pd.DataFrame(data = zip(score[0],score[1],score[2],score[3],score[4],score[5]),\
                        columns = sheet_names,index =  list_year) # creating data frame

    

    if write==True:
        txtfile = open(filename,'w+',encoding='utf-8') # opening text file
        txtfile.write(data.to_string()) # writing the dataframe in the text file
        txtfile.close # closing the file
        return None

    else:
        return data 


def exec(list_req, wb,write,Total):
    """
     excute Command for either writing the dataframe into txt or just put them in a big dataframe
     ==============================================================================================================

     Input:
     ------
     list_req = list of tickers required
     web =  loaded excel workbook
     write: boolean True if txt files are required and None return
                    False if dataframe form required and returned

    Total: boolean True: return one big Dataframe
                   False: returna  list of Dataframes, corresponding to the order of the list_req's tickers
    """
    ret = None

    if write==True:
        for tick in list_req:
            retrieve_score(wb,tick,increase=True,write = write)
            retrieve_score(wb,tick,increase=False,write = write) 
    
    else:
        if Total == True:
            ret_inc = retrieve_score(wb,list_req[0],increase=True,write = write)
            ret_score = retrieve_score(wb,list_req[0],increase=False,write = write)
            for tick in list_req[1:]:
                ret_inc = ret_inc.append(retrieve_score(wb,tick,increase=True,write = write))
                ret_score = ret_score.append(retrieve_score(wb,tick,increase=False,write = write))
        
        else:
            ret_inc = []
            ret_score = []
            for tick in list_req[1:]:
                ret_inc.append(retrieve_score(wb,tick,increase=True,write = write))
                ret_score.append(retrieve_score(wb,tick,increase=False,write = write))


        ret = (ret_score,ret_inc)

    
    return ret



    


