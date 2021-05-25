from openpyxl import load_workbook


def searchcatsub(sr,D):
    """ This function return a list of subcategory's name for a category or the categor's name for the topic
    input:
    sr: Excel Search Engine for a specific Sheet
    D: the Name of the column from which we want to extract the names

    Output:
    list containing the list of lists containing the names of sub categories
    """
    result = [] # initializing
    for j in range(1,len(sr[D])+1): # looping through all rows

        k = D + str(j) # creating the cell's ID

        if sr[k].value is None: # if the va 
            continue # skips
        
        else: # if not null
            result.append(sr[k].value.lower()) # adopting a norm of lower values
    
    return result

def clean(arr): 
    """
    clean array from empty '' caused by the splitting
    """
    cleaned = [] 
    for i in arr:

        if len(i) == 0: # such strings have length 0
            continue # so we skip them

        else:
            cleaned.append(i) # we append them otherwise

    return cleaned

def searchvalues(sr,p,n):
    """
        this function has a job of collecting the words tokenize them and add them to list, which also will be categorised
        in a list according the the category/subcategory 
    
        Input:
        
        sr: Excel Search engine given a sheet
        p: the tag of the Column where there is the sub-category/category
        n: the tag of the column where there is the actual words corresponding

        Output:
        results: list of lists, the lists inside are divided to accomodate for the number of subsequent categories/sub
    """
    
    result , sub = [] , []

    for j in range(1,len(sr[p])+1):

        prev = p + str(j) # creating the cell's ID of the previous column ie. the sub/categories

        act = n + str(j)  # creating the cell's ID of the cell containing the values

        if sr[act].value is None: # if the cell of words is empty, skip this cell
            
            continue

        if sr[prev].value is not None: # if the prev cell is not empty means we have a new sub/cat
            
            result.append(clean(sub)) # append the previously collected , cleaned
            
            sub = [] # reinitiate the same list 

        if "," in sr[act].value: # if the cell contains many words
            
            # lower case, eliminate spaces and split at ',' and add this list to sub
            
            sub+=sr[act].value.lower().replace(" ","").split(",") 
        
        else:

            sub.append(sr[act].value.lower()) # otherwise it appends the lower case value
    
    result.append(clean(sub))  # appending the last result as its was not appended

    return result[1:] # the first list initialted is added directly so it was taken out
        


def importing(sr): 
    
    """ 
        this function collects the data collected in the previous functions and returns a dictionary with multiple layers
        with keys equal to categories sub categories and if the the category has a some words its added under key : 'self'

        Input:
        sr: Excel Engine related to a specific excel sheet

        Output:
        Result: dict, containing all the words under their correct distribution in the excel file

    """

    # first getting the Data
    Topic_value = searchvalues(sr,'A','B') # the words associated directly with the topic
    Categories_name = searchcatsub(sr,'C') # the list of categories under the topic
    Categories_values = searchvalues(sr,'C','D') # the words assigned to each of these categories
    Subcateg_name = searchcatsub(sr,'E') # the list of sub category under each cateogry
    Subcateg_values = searchvalues(sr,'E','F') # list of words for each of these sub categories
    inhertence = searchvalues(sr,'C','E') # the list of sub cateogies with respect to each category
    
    # backward induction to build the dictionary

    last_layer = {} # initializing the last layer of our dictionary
    
    for i in range(len(Subcateg_name)): # loopoing through each subcategory name

        # appending to the last layer keys being the names of subcategory and the 
        # value being the list of corresponding words 

        last_layer.update({Subcateg_name[i]:Subcateg_values[i]}) 

    second_layer ={} #initializing the second to last layer

    for j in range(len(Categories_name)): # looping through all categories name

        second_layerh = {} # initalizing the hidden layer i.e. the dictionary in the dictionary

        second_layerh.update({'self' : Categories_values[j]}) # appending first the own words of the category
        
        for k in inhertence[j]: # then looping through all its inheritance ie. the subs corresponding the each category 

            # for each inheritant sub category, adding the name as key and the value 
            # as the previous layer with the same key

            second_layerh.update({k : last_layer[k]}) 

        second_layer.update({Categories_name[j]:second_layerh}) # then adding all this hidden dictionary in the second layer one
    
    Result = {} # this is the return dictionary containint all the words neatly oraganized by category and subs

    Result.update({'self':Topic_value[0]}) # the topic has its words, adding them under the key self

    for l in Categories_name: # looping through all categories
        
        # appending the dictionary with key as the cateogry and the value as the dictionary of the category
        Result.update({l:second_layer[l]}) 
        

    return Result

########################################################################################################################################################################
# the first one was easy to deal with its the second one which is really a mess

wb = load_workbook(filename="LoughranMcDonald_SentimentWordLists_2018.xlsx") # initialzing for the first Sentiment T

sheetn = wb.sheetnames[1:] # the first sheet is just information we don't need

Sentinents = {} # initiliazing the dictionar

for i in range(len(sheetn)): # looping through all sheets

    listword=[] # in each sheet we redifine a new list

    sr=wb[sheetn[i]] # we set the engine to work in the specific sheet

    for j in range(1,sr.max_row+1): # we loop till the last row

        k = 'A' + str(j) # creating the the ID of the cell
        
        listword.append(sr[k].value.lower()) # appending its lower case value
    
    Sentinents.update({sheetn[i]:listword}) # we append to the main dictionary the list collected, with a key equal to the name of the sheet


#########################################################################################################################################################################